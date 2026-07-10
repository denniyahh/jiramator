"""Spreadsheet ingestion for CSV and XLSX import workflows."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

from jiramator.encoding import detect_encoding


def _normalize_headers(headers: list[object]) -> list[str]:
    return [str(header).strip() for header in headers]


def _coerce_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _read_csv(
    path: Path,
    *,
    max_rows: int | None = None,
    encoding_override: str | None = None,
) -> list[dict[str, str]]:
    enc = detect_encoding(path, override=encoding_override)
    if enc != "utf-8":
        # DC-1: stdout stays clean; user-facing notes go to stderr. Use
        # rich.Console here for consistency with the rest of the CLI's
        # styled output, though the message itself is plain text.
        from rich.console import Console
        # soft_wrap so a long CSV path isn't hard-wrapped across lines.
        Console(stderr=True).print(f"Read {path} as {enc}", soft_wrap=True)

    with path.open(newline="", encoding=enc) as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            return []

        headers = _normalize_headers(list(reader.fieldnames))
        rows: list[dict[str, str]] = []
        for index, raw_row in enumerate(reader):
            if max_rows is not None and index >= max_rows:
                break
            normalized: dict[str, str] = {}
            for original, normalized_header in zip(reader.fieldnames, headers, strict=True):
                normalized[normalized_header] = _coerce_cell(raw_row.get(original))
            rows.append(normalized)
        return rows


def _read_xlsx(
    path: Path,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = None,
) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]

        rows_iter = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            return []

        headers = _normalize_headers(list(raw_headers))
        rows: list[dict[str, str]] = []
        for row in rows_iter:
            normalized = {
                header: _coerce_cell(value)
                for header, value in zip(headers, row, strict=False)
            }
            if len(row) < len(headers):
                for header in headers[len(row):]:
                    normalized[header] = ""
            if all(value == "" for value in normalized.values()):
                continue
            rows.append(normalized)
            if max_rows is not None and len(rows) >= max_rows:
                break
        return rows
    finally:
        workbook.close()


def read_spreadsheet(
    path: str | Path,
    *,
    sheet_name: str | None = None,
    max_rows: int | None = None,
    encoding_override: str | None = None,
) -> list[dict[str, str]]:
    """Read a CSV or XLSX spreadsheet into a list of row dicts.

    Args:
        path: CSV or XLSX file path.
        sheet_name: For XLSX only — sheet name (defaults to first sheet).
        max_rows: Optional row limit.
        encoding_override: For CSV only — bypass encoding detection and
            use this codec verbatim. Ignored for XLSX (binary reader).
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return _read_csv(
            path, max_rows=max_rows, encoding_override=encoding_override,
        )
    if suffix == ".xlsx":
        return _read_xlsx(path, sheet_name=sheet_name, max_rows=max_rows)
    if suffix == ".xls":
        raise ValueError(".xls is not supported; save the spreadsheet as .xlsx first")
    raise ValueError(f"Unsupported spreadsheet file type: {path.suffix}")
