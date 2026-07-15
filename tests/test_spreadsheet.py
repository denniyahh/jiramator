"""Tests for CSV/XLSX spreadsheet ingestion used by the import command."""

from __future__ import annotations

from pathlib import Path

import pytest


class TestReadSpreadsheet:
    def test_read_csv_returns_rows(self, tmp_path: Path):
        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.csv"
        path.write_text("Summary,Priority\nRisk A,High\nRisk B,Medium\n")

        rows = read_spreadsheet(path)

        assert rows == [
            {"Summary": "Risk A", "Priority": "High"},
            {"Summary": "Risk B", "Priority": "Medium"},
        ]

    def test_read_csv_trims_header_whitespace(self, tmp_path: Path):
        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.csv"
        path.write_text(" Summary , Priority \nRisk A,High\n")

        rows = read_spreadsheet(path)

        assert rows == [{"Summary": "Risk A", "Priority": "High"}]

    def test_read_xlsx_returns_rows(self, tmp_path: Path):
        import openpyxl

        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Risks"
        ws.append(["Summary", "Priority"])
        ws.append(["Risk A", "High"])
        ws.append(["Risk B", "Medium"])
        wb.save(path)

        rows = read_spreadsheet(path)

        assert rows == [
            {"Summary": "Risk A", "Priority": "High"},
            {"Summary": "Risk B", "Priority": "Medium"},
        ]

    def test_read_xlsx_can_select_sheet(self, tmp_path: Path):
        import openpyxl

        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.xlsx"
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "IgnoreMe"
        ws1.append(["Summary", "Priority"])
        ws1.append(["Wrong", "Low"])
        ws2 = wb.create_sheet("Risks")
        ws2.append(["Summary", "Priority"])
        ws2.append(["Risk A", "High"])
        wb.save(path)

        rows = read_spreadsheet(path, sheet_name="Risks")

        assert rows == [{"Summary": "Risk A", "Priority": "High"}]

    def test_none_cells_become_empty_strings(self, tmp_path: Path):
        import openpyxl

        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Summary", "Priority"])
        ws.append(["Risk A", None])
        wb.save(path)

        rows = read_spreadsheet(path)

        assert rows == [{"Summary": "Risk A", "Priority": ""}]

    def test_whole_number_float_cells_render_without_decimal(self, tmp_path: Path):
        """Excel/openpyxl often stores whole numbers as floats internally
        (e.g. a cell showing "1" reads back as 1.0). Left unhandled, this
        breaks exact-match downstream lookups like `value_aliases`, which
        expect "1" not "1.0"."""
        import openpyxl

        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Summary", "Code Complexity"])
        ws.append(["Risk A", 1.0])
        ws.append(["Risk B", 2.5])
        wb.save(path)

        rows = read_spreadsheet(path)

        assert rows == [
            {"Summary": "Risk A", "Code Complexity": "1"},
            {"Summary": "Risk B", "Code Complexity": "2.5"},
        ]

    def test_max_rows_limits_output(self, tmp_path: Path):
        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.csv"
        path.write_text("Summary\nRisk A\nRisk B\nRisk C\n")

        rows = read_spreadsheet(path, max_rows=2)

        assert rows == [{"Summary": "Risk A"}, {"Summary": "Risk B"}]

    def test_skips_fully_empty_rows(self, tmp_path: Path):
        import openpyxl

        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Summary", "Priority"])
        ws.append(["Risk A", "High"])
        ws.append([None, None])
        ws.append(["Risk B", "Low"])
        wb.save(path)

        rows = read_spreadsheet(path)

        assert rows == [
            {"Summary": "Risk A", "Priority": "High"},
            {"Summary": "Risk B", "Priority": "Low"},
        ]

    def test_rejects_xls_extension(self, tmp_path: Path):
        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "legacy.xls"
        path.write_text("not really xls")

        with pytest.raises(ValueError, match=".xls is not supported"):
            read_spreadsheet(path)

    def test_rejects_unsupported_extension(self, tmp_path: Path):
        from jiramator.spreadsheet import read_spreadsheet

        path = tmp_path / "sample.txt"
        path.write_text("hello")

        with pytest.raises(ValueError, match="Unsupported spreadsheet file type"):
            read_spreadsheet(path)


# ---------------------------------------------------------------------------
# Encoding detection wiring (Plan 01-02 Task 2)
# ---------------------------------------------------------------------------


FIXTURE_DIR = Path(__file__).parent / "fixtures" / "csv_encodings"


class TestEncodingDetection:
    """Tests for encoding_override parameter and detect_encoding wiring."""

    # --- R1, R2: regression / preservation ---

    def test_R1_plain_utf8_returns_expected_rows(self):
        """R1: read_spreadsheet on plain utf8.csv returns the same row dicts
        as before the encoding-detection wiring."""
        from jiramator.spreadsheet import read_spreadsheet
        rows = read_spreadsheet(FIXTURE_DIR / "utf8.csv")
        assert rows == [{"name": "foo", "value": "1"}]

    def test_R2_xlsx_path_unaffected_by_encoding_override(self, tmp_path):
        """R2: xlsx ignores encoding_override (binary reader; openpyxl owns
        encoding). Passing the parameter must not raise or alter results."""
        import openpyxl
        from jiramator.spreadsheet import read_spreadsheet

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["foo", 1])
        path = tmp_path / "data.xlsx"
        wb.save(path)

        rows_default = read_spreadsheet(path)
        rows_with_override = read_spreadsheet(path, encoding_override="cp1252")
        assert rows_default == rows_with_override
        assert rows_default == [{"name": "foo", "value": "1"}]

    # --- N1: utf-8 with BOM ---

    def test_N1_utf8_sig_strips_bom_from_header(self):
        """N1: utf-8-sig codec strips the leading BOM during decode so the
        header is 'name', not '\\ufeffname' (T-01-09 mitigation)."""
        from jiramator.spreadsheet import read_spreadsheet
        rows = read_spreadsheet(FIXTURE_DIR / "utf8_sig.csv")
        assert rows == [{"name": "foo", "value": "1"}]
        # No BOM character in any header key
        for row in rows:
            for key in row:
                assert "\ufeff" not in key

    # --- N2: cp1252 round-trip preserves Unicode codepoints ---

    def test_N2_cp1252_preserves_unicode_codepoints(self):
        """N2: cp1252 file with smart quotes / em-dash / £ decodes to the
        proper Unicode codepoints (no mojibake, no raw bytes)."""
        from jiramator.spreadsheet import read_spreadsheet
        rows = read_spreadsheet(FIXTURE_DIR / "cp1252.csv")
        assert len(rows) == 1
        row = rows[0]
        # Keys
        assert set(row.keys()) == {"name", "note"}
        # 'O\u2019Brien' — curly apostrophe preserved
        assert row["name"] == "O\u2019Brien"
        # '\u201csmart\u201d \u2014 \u00a3100' — smart quotes, em-dash, £
        assert "\u201c" in row["note"] or "\u201d" in row["note"]
        assert "\u2014" in row["note"]
        assert "\u00a3" in row["note"] or "100" in row["note"]

    # --- N3: utf-16-le with BOM ---

    def test_N3_utf16_le_bom_decodes_correctly(self):
        from jiramator.spreadsheet import read_spreadsheet
        rows = read_spreadsheet(FIXTURE_DIR / "utf16_le_bom.csv")
        assert rows == [{"name": "foo", "value": "1"}]

    # --- N4: encoding_override bypasses charset-normalizer ---

    def test_N4_override_bypasses_charset_normalizer(self, monkeypatch):
        """N4: When encoding_override is given, charset_normalizer.from_bytes
        is never invoked. Patch it to fail loudly; if the test still passes,
        the override branch took precedence."""
        from jiramator.spreadsheet import read_spreadsheet

        def _explode(*_a, **_kw):
            pytest.fail("charset_normalizer.from_bytes called despite override")

        monkeypatch.setattr("charset_normalizer.from_bytes", _explode)
        rows = read_spreadsheet(
            FIXTURE_DIR / "cp1252.csv", encoding_override="cp1252"
        )
        assert len(rows) == 1
        assert rows[0]["name"] == "O\u2019Brien"

    # --- N5: stderr announcement only on non-UTF-8 ---

    def test_N5a_announcement_on_cp1252(self, capsys):
        """N5a: cp1252 import emits 'Read <path> as <encoding>' to stderr."""
        import re
        from jiramator.spreadsheet import read_spreadsheet
        read_spreadsheet(FIXTURE_DIR / "cp1252.csv")
        captured = capsys.readouterr()
        # Match shape; exact encoding name varies (Pitfall 5).
        pattern = re.compile(
            r"Read .* as (cp1252|cp1250|windows-1252|windows-1250|"
            r"iso-8859-1|iso-8859-15|latin-1|latin_1)"
        )
        assert pattern.search(captured.err), (
            f"stderr did not match expected announcement: {captured.err!r}"
        )

    def test_N5b_no_announcement_on_plain_utf8(self, capsys):
        """N5b: plain UTF-8 import is silent on stderr (happy path noise-free)."""
        from jiramator.spreadsheet import read_spreadsheet
        read_spreadsheet(FIXTURE_DIR / "utf8.csv")
        captured = capsys.readouterr()
        assert "Read" not in captured.err
        assert captured.err == "" or captured.err.isspace()

    def test_N5c_announces_utf8_sig_encoding(self, capsys):
        """N5c: utf-8-sig is distinct from utf-8; the announcement IS emitted
        so the user can see what encoding was detected and override if needed.
        Decision: announce for any non-utf-8 encoding (including utf-8-sig
        and utf-16). This preserves transparency over silence."""
        from jiramator.spreadsheet import read_spreadsheet
        read_spreadsheet(FIXTURE_DIR / "utf8_sig.csv")
        captured = capsys.readouterr()
        # utf-8-sig != utf-8, so an announcement IS expected here.
        assert "Read" in captured.err
        assert "utf-8-sig" in captured.err

    # --- N6: undetectable encoding surfaces actionable error ---

    def test_N6_undetectable_raises_with_actionable_hint(self, tmp_path, monkeypatch):
        """N6: When detection fails, the surfaced ValueError names both
        cp1252 and utf-16-le as suggested overrides. We force the failure
        path by patching charset_normalizer.from_bytes to return a results
        object whose .best() is None — the only way to deterministically
        hit the ValueError branch (random binary may produce a guess)."""
        from jiramator.spreadsheet import read_spreadsheet

        class _NoBest:
            def best(self):
                return None

        monkeypatch.setattr("charset_normalizer.from_bytes", lambda *_a, **_kw: _NoBest())

        # Bytes that fail strict UTF-8 decode and have no BOM.
        p = tmp_path / "weird.csv"
        p.write_bytes(b"\xff\xfe\xfdGARBAGE\x80\x81\x82")  # not a valid BOM combo
        # NOTE: \xff\xfe IS the UTF-16-LE BOM. Avoid it — use bytes that
        # truly bypass the BOM table.
        p.write_bytes(b"\x80\x81\x82\x83 some bytes \x90\x91")

        with pytest.raises(ValueError, match=r"Could not detect encoding") as exc_info:
            read_spreadsheet(p)
        msg = str(exc_info.value)
        assert "cp1252" in msg
        assert "utf-16-le" in msg
